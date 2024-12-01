from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles import PatternFill


class ManagerXLSX:

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

        self.ws.merge_cells('A1:F1')
        self.ws['A1'] = "ПЕРЕЧЕТНАЯ ВЕДОМОСТЬ ДЕРЕВЬЕВ И КУСТАРНИКОВ"
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['A1'].font = Font(bold=True, size=14)

    def create(self, data: list) -> Workbook:
        headers = self.create_headers()
        self.set_styles(headers)
        self.add_data(data)
        self.set_width()
        self.set_height()
        self.configure_data_style_cells(headers)
        self.set_colors_headers()

        return self.wb

    def create_headers(self) -> list:
        headers: list = [
            "№ п/п", "Наименование пород", "Кол-во в шт.", "Диаметр, Р, см",
            "Высота, м", "Характеристика состояния зеленых насаждений"
        ]
        self.ws.append(headers)

        return headers

    def set_styles(self, headers: list):
        for col in range(1, len(headers) + 1):
            cell = self.ws.cell(row=2, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    def add_data(self, data: list):
        for row in data:
            self.ws.append(row)

    def set_width(self):
        col_widths = [8, 25, 15, 15, 10, 40]
        for i, width in enumerate(col_widths, start=1):
            self.ws.column_dimensions[chr(64 + i)].width = width

    def set_height(self):
        self.ws.row_dimensions[1].height = 30
        self.ws.row_dimensions[2].height = 25

    def configure_data_style_cells(self, headers: list):
        for row in self.ws.iter_rows(min_row=3, max_row=self.ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

    def set_colors_headers(self):
        fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for row in [1, 2]:
            for cell in self.ws[row]:
                cell.fill = fill
